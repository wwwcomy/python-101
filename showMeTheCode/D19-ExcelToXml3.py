from openpyxl import load_workbook
import xml.etree.ElementTree as ET

#number.xlsx is the out put of D16
wb = load_workbook(filename = 'numbers.xlsx')
sheet_ranges = wb['Sheet']
dict = []
print("sheet_ranges -> " + str(sheet_ranges.max_row))
for i in range (0, sheet_ranges.max_row):
    dict.append([])
    for val in sheet_ranges[i+1]:
        dict[i].append(val.value)
print(dict)

# 保存修改了的xml
root = ET.Element("root")
numbers = ET.SubElement(root, "numbers")
numbers.insert(0,ET.Comment('''
	数字信息
	'''))
numbers.text = str(dict)
tree = ET.ElementTree(root)
tree.write("numbers.xml",encoding='UTF8')